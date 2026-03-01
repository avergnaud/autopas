"""Sélection des fichiers de référence du corpus par score de similarité."""
from __future__ import annotations

import json
import logging
from pathlib import Path

from app.config import BASE_DIR

logger = logging.getLogger(__name__)

CORPUS_DIR = BASE_DIR / "data" / "corpus" / "files"
_MAX_REF_CHARS = 30_000  # ~30 KB per reference file sent to Claude


def _load_metadata(json_path: Path) -> dict | None:
    try:
        return json.loads(json_path.read_text(encoding="utf-8"))
    except Exception as exc:
        logger.warning("Failed to load corpus metadata %s: %s", json_path, exc)
        return None


def _score(cadrage: dict, meta: dict) -> int:
    """Compute integer similarity score between cadrage answers and corpus metadata."""
    score = 0

    # Type de prestation (+3)
    if (
        str(cadrage.get("type_prestation", "")).lower()
        == str(meta.get("type_prestation", "")).lower()
        and cadrage.get("type_prestation")
    ):
        score += 3

    # Hébergement données (+2)
    if (
        str(cadrage.get("hebergement_donnees", "")).lower()
        == str(meta.get("hebergement_donnees", "")).lower()
        and cadrage.get("hebergement_donnees")
    ):
        score += 2

    # Activités — intersection (+2)
    cadrage_acts = {str(a).lower() for a in cadrage.get("activites", []) if a}
    meta_acts = {str(a).lower() for a in meta.get("activites", []) if a}
    if cadrage_acts & meta_acts:
        score += 2

    # Sous-traitance RGPD (+2)
    c_rgpd = cadrage.get("sous_traitance_rgpd")
    m_rgpd = meta.get("sous_traitance_rgpd")
    if c_rgpd is not None and m_rgpd is not None and c_rgpd == m_rgpd:
        score += 2

    # Poste de travail (+1)
    if (
        str(cadrage.get("poste_travail", "")).lower()
        == str(meta.get("poste_travail", "")).lower()
        and cadrage.get("poste_travail")
    ):
        score += 1

    # Expertise Atlassian (+1)
    c_atl = cadrage.get("expertise_atlassian")
    m_atl = meta.get("expertise_atlassian")
    if c_atl is not None and m_atl is not None and c_atl == m_atl:
        score += 1

    # Lieu de travail — intersection (+1)
    c_lieu = {str(l).lower() for l in cadrage.get("lieu_travail", []) if l}
    m_lieu = {str(l).lower() for l in meta.get("lieu_travail", []) if l}
    if c_lieu & m_lieu:
        score += 1

    # Format questionnaire (+1)
    if (
        str(cadrage.get("format", "")).lower() == str(meta.get("format", "")).lower()
        and cadrage.get("format")
    ):
        score += 1

    return score


def _read_reference_content(filepath: Path) -> str:
    """Read reference file content as flat text, capped at _MAX_REF_CHARS."""
    try:
        ext = filepath.suffix.lower()
        if ext == ".xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            lines = []
            total = 0
            truncated = False
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                lines.append(f"=== {sheet_name} ===")
                for row in ws.iter_rows(values_only=True):
                    row_text = " | ".join(str(c) if c else "" for c in row)
                    if row_text.strip():
                        lines.append(row_text)
                        total += len(row_text)
                        if total >= _MAX_REF_CHARS:
                            truncated = True
                            break
                if truncated:
                    break
            wb.close()
            content = "\n".join(lines)
            if truncated:
                content += "\n[... tronqué ...]"
            return content
        elif ext == ".docx":
            from docx import Document as DocxDocument
            doc = DocxDocument(str(filepath))
            content = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        elif ext in (".txt", ".md"):
            content = filepath.read_text(encoding="utf-8")
        else:
            return ""
        if len(content) > _MAX_REF_CHARS:
            content = content[:_MAX_REF_CHARS] + "\n[... tronqué ...]"
        return content
    except Exception as exc:
        logger.warning("Failed to read reference file %s: %s", filepath, exc)
    return ""


def select_references(
    cadrage: dict, max_files: int = 3
) -> list[tuple[str, str, dict]]:
    """
    Return the top max_files reference files scored against cadrage.

    Returns:
        List of (filepath_str, content_text, metadata_dict), best match first.
    """
    if not CORPUS_DIR.exists():
        logger.warning("Corpus directory not found: %s", CORPUS_DIR)
        return []

    scored: list[tuple[int, Path, dict]] = []
    for json_path in CORPUS_DIR.glob("*.json"):
        meta = _load_metadata(json_path)
        if not meta:
            continue
        filename = meta.get("filename", "")
        file_path = CORPUS_DIR / filename
        if not file_path.exists():
            logger.debug("Corpus file missing: %s", file_path)
            continue
        scored.append((_score(cadrage, meta), file_path, meta))

    scored.sort(key=lambda x: x[0], reverse=True)

    result = []
    for score, fpath, meta in scored[:max_files]:
        logger.info("Reference selected: %s (score=%d)", fpath.name, score)
        content = _read_reference_content(fpath)
        result.append((str(fpath), content, meta))

    return result
