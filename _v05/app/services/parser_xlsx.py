"""Parsing and writing xlsx questionnaires with openpyxl."""
from __future__ import annotations

import logging
import shutil
import zipfile
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string

from app.models.project import DocumentStructure
from app.services.anonymizer import Anonymizer

logger = logging.getLogger(__name__)
_MAX_EMPTY_ROWS = 20  # stop scanning after this many consecutive empty question cells


def extract_raw_content(filepath: Path, max_rows: int = 25) -> str:
    """Extract first max_rows rows from each sheet as text for Claude structure analysis."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    lines = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"=== Onglet : {sheet_name} ===")
        for i, row in enumerate(ws.iter_rows(max_row=max_rows, values_only=True), 1):
            cells = [str(c) if c is not None else "" for c in row]
            # Skip completely empty rows
            if any(c.strip() for c in cells):
                lines.append(f"Ligne {i} : {' | '.join(cells)}")
        lines.append("")
    wb.close()
    return "\n".join(lines)


def extract_questions(filepath: Path, structure: DocumentStructure) -> list[dict]:
    """Extract all questions from the xlsx. Returns list of {id, question, sheet, row}."""
    if not structure.sheets:
        return []

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    questions = []

    for sheet_info in structure.sheets:
        if not sheet_info.has_questions:
            continue
        if sheet_info.name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found in workbook", sheet_info.name)
            continue

        ws = wb[sheet_info.name]
        q_col_idx = column_index_from_string(sheet_info.question_column)
        id_col_idx = (
            column_index_from_string(sheet_info.id_column)
            if sheet_info.id_column
            else None
        )

        # Use iter_rows() — the correct streaming API for read_only mode.
        # ws.cell() + range(max_row) is O(n²) in read_only and can loop over
        # 1 048 576 rows if the xlsx dimension tag is stale.
        empty_streak = 0
        for row_idx, row_cells in enumerate(
            ws.iter_rows(min_row=sheet_info.first_data_row)
        ):
            row_num = sheet_info.first_data_row + row_idx
            q_val = row_cells[q_col_idx - 1].value if q_col_idx - 1 < len(row_cells) else None
            q_text = str(q_val).strip() if q_val is not None else ""

            if not q_text:
                empty_streak += 1
                if empty_streak >= _MAX_EMPTY_ROWS:
                    break
                continue
            empty_streak = 0

            q_id = ""
            if id_col_idx and id_col_idx - 1 < len(row_cells):
                id_val = row_cells[id_col_idx - 1].value
                q_id = str(id_val).strip() if id_val is not None else ""
            if not q_id:
                q_id = f"{sheet_info.name}_row{row_num}"

            questions.append({
                "id": q_id,
                "question": q_text,
                "sheet": sheet_info.name,
                "row": row_num,
            })

    wb.close()
    return questions


def format_questionnaire_for_claude(questions: list[dict]) -> str:
    """Format extracted questions as text block for the Claude prompt."""
    lines = []
    for q in questions:
        lines.append(f"ID: {q['id']}")
        lines.append(f"Question: {q['question']}")
        lines.append("")
    return "\n".join(lines)


def write_responses(
    filepath: Path, responses: list[dict], structure: DocumentStructure
) -> None:
    """Write Claude responses into response columns of the xlsx file."""
    if not structure.sheets:
        return

    response_map = {r["question_id"]: r["response"] for r in responses}

    wb = openpyxl.load_workbook(filepath)

    for sheet_info in structure.sheets:
        if not sheet_info.has_questions or sheet_info.name not in wb.sheetnames:
            continue

        ws = wb[sheet_info.name]
        q_col_idx = column_index_from_string(sheet_info.question_column)
        id_col_idx = (
            column_index_from_string(sheet_info.id_column)
            if sheet_info.id_column
            else None
        )
        resp_col_idxs = [
            column_index_from_string(c) for c in sheet_info.response_columns
        ]

        if not resp_col_idxs:
            logger.warning("No response columns for sheet '%s'", sheet_info.name)
            continue

        for row_num in range(sheet_info.first_data_row, ws.max_row + 1):
            # Determine question ID
            q_id = ""
            if id_col_idx:
                id_cell = ws.cell(row=row_num, column=id_col_idx)
                q_id = str(id_cell.value).strip() if id_cell.value else ""
            if not q_id:
                q_cell = ws.cell(row=row_num, column=q_col_idx)
                if q_cell.value:
                    q_id = f"{sheet_info.name}_row{row_num}"

            if q_id in response_map:
                ws.cell(row=row_num, column=resp_col_idxs[0]).value = response_map[q_id]

    wb.save(filepath)
    wb.close()


def delete_unlisted_sheets(filepath: Path, sheet_names_to_keep: list[str]) -> None:
    """Delete from the xlsx file every sheet whose name is NOT in sheet_names_to_keep."""
    if not sheet_names_to_keep:
        return
    wb = openpyxl.load_workbook(filepath)
    to_delete = [name for name in wb.sheetnames if name not in sheet_names_to_keep]
    for name in to_delete:
        logger.info("Deleting sheet '%s' from %s", name, filepath.name)
        del wb[name]
    if to_delete:
        wb.save(filepath)
    wb.close()


def apply_anonymization(filepath: Path, anonymizer: Anonymizer) -> None:
    """Apply anonymization to xlsx by modifying ZIP/XML content directly (memory-efficient)."""
    _xlsx_zip_replace(filepath, anonymizer.anonymize)


def apply_deanonymization(filepath: Path, anonymizer: Anonymizer) -> None:
    """Apply deanonymization to xlsx by modifying ZIP/XML content directly (memory-efficient)."""
    _xlsx_zip_replace(filepath, anonymizer.deanonymize)


def _xlsx_zip_replace(filepath: Path, transform_fn) -> None:
    """Memory-efficient text replacement inside an xlsx (ZIP) by processing XML entries.

    Targets xl/sharedStrings.xml (where most text is stored) and worksheet XML files
    (for inline strings). Avoids loading a full openpyxl workbook in write mode.
    """
    tmp = filepath.with_suffix(".~tmp.xlsx")
    try:
        with (
            zipfile.ZipFile(filepath, "r") as zin,
            zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as zout,
        ):
            for info in zin.infolist():
                data = zin.read(info.filename)
                name = info.filename
                if name == "xl/sharedStrings.xml" or (
                    name.startswith("xl/worksheets/") and name.endswith(".xml")
                ):
                    try:
                        text = data.decode("utf-8")
                        text = transform_fn(text)
                        data = text.encode("utf-8")
                    except Exception as exc:
                        logger.warning("ZIP replace failed for %s: %s", name, exc)
                zout.writestr(info, data)
        shutil.move(str(tmp), str(filepath))
    except Exception:
        try:
            tmp.unlink(missing_ok=True)
        except Exception:
            pass
        raise
