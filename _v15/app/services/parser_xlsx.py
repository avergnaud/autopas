"""PAS Assistant — xlsx questionnaire parser.

Handles reading questions from an anonymized xlsx and writing Claude-generated
responses back into a copy of that file.
"""

import logging
import re
import shutil
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.workbook.defined_name import DefinedName

from app.services.anonymizer import safe_local_defined_names

_XLSX_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

logger = logging.getLogger(__name__)


def _merged_header_rows(ws, col_r_idx: int) -> set[int]:
    """Return row indices where col_r_idx is a MergedCell slave (top-left is in another col).

    These rows are section/category header rows whose label spans across the response
    column.  Writing into them would overwrite the header text, so they must be skipped.

    Args:
        ws: openpyxl Worksheet (must NOT be opened in read_only mode).
        col_r_idx: 1-based index of the response column.

    Returns:
        Set of 1-based row indices to skip.
    """
    skip: set[int] = set()
    for mr in ws.merged_cells.ranges:
        # The merge range covers the response column but starts to its left
        if mr.min_col < col_r_idx <= mr.max_col:
            for r in range(mr.min_row, mr.max_row + 1):
                skip.add(r)
    return skip


def read_questions(xlsx_path: Path, structure: dict) -> list[dict]:
    """Extract questions from an xlsx file using confirmed structure.

    Uses a two-pass strategy to stay memory-efficient:
    - Pass 1: brief full-mode load to collect merged cell ranges, then immediately freed.
    - Pass 2: read_only streaming load for actual question extraction.

    Args:
        xlsx_path: Path to the xlsx file (anonymized).
        structure: Structure dict with keys: selected_sheet, first_data_row,
                   col_id (optional), col_question, col_response (optional).

    Returns:
        List of dicts with keys: row (int), question_id (str), question_text (str).
        Rows where question_text is empty are skipped.
        Rows where col_response is a MergedCell slave (section header rows) are skipped.
    """
    sheet_name = structure.get("selected_sheet")
    first_data_row = int(structure.get("first_data_row") or 2)
    col_id_letter = structure.get("col_id")
    col_q_letter = structure.get("col_question")
    col_r_letter = structure.get("col_response")

    if not col_q_letter:
        raise ValueError("Structure manquante : col_question est obligatoire.")

    col_q_idx = column_index_from_string(col_q_letter.upper())  # 1-based
    col_id_idx = column_index_from_string(col_id_letter.upper()) if col_id_letter else None
    col_r_idx = column_index_from_string(col_r_letter.upper()) if col_r_letter else None

    # Pass 1 — brief full-mode load just to collect merged cell ranges, then freed immediately.
    # read_only mode does not reliably expose ws.merged_cells.ranges across all openpyxl versions.
    header_rows: set[int] = set()
    if col_r_idx:
        _wb = load_workbook(xlsx_path, data_only=True)
        _ws = _wb[sheet_name] if sheet_name and sheet_name in _wb.sheetnames else _wb.active
        header_rows = _merged_header_rows(_ws, col_r_idx)
        _wb.close()
        del _wb, _ws  # release memory before the streaming pass
        if header_rows:
            logger.info(
                "read_questions: %d section-header rows will be skipped (merged into col_response): %s",
                len(header_rows),
                sorted(header_rows),
            )

    # Pass 2 — read_only streaming load for memory-efficient question extraction.
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    questions: list[dict] = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx < first_data_row:
            continue
        if row_idx in header_rows:
            continue

        # Extract question text
        q_text_raw = row[col_q_idx - 1] if col_q_idx - 1 < len(row) else None
        q_text = str(q_text_raw).strip() if q_text_raw is not None else ""
        if not q_text:
            continue

        # Extract question id
        if col_id_idx is not None and col_id_idx - 1 < len(row):
            q_id_raw = row[col_id_idx - 1]
            q_id = str(q_id_raw).strip() if q_id_raw is not None else str(row_idx)
        else:
            q_id = str(row_idx)

        questions.append({"row": row_idx, "question_id": q_id, "question_text": q_text})

    wb.close()
    logger.info("read_questions: extracted %d questions from %s", len(questions), xlsx_path.name)
    return questions


def _resolve_named_range_from_xml(xlsx_path: Path, name: str, sheet_idx: int | None) -> str | None:
    """Resolve a named range from raw workbook XML, handling local sheet scope.

    openpyxl loses local defined names when a same-named workbook-level #REF! entry
    exists.  This reads the raw XML directly to find the correct range reference.

    Prefers sheet-local definition (matching sheet_idx) over workbook-level fallback.

    Args:
        xlsx_path: Path to the xlsx file.
        name: Name of the defined name to resolve.
        sheet_idx: 0-based index of the sheet (localSheetId) to prefer, or None.

    Returns:
        Range reference string (e.g. ``'Sheet1'!$A$1:$A$4``) or None if not found.
    """
    try:
        with zipfile.ZipFile(xlsx_path) as z:
            with z.open("xl/workbook.xml") as f:
                tree = ET.parse(f)
    except Exception:
        return None

    local_match: str | None = None
    workbook_match: str | None = None

    for dn in tree.iter(f"{{{_XLSX_NS}}}definedName"):
        if dn.get("name", "") != name:
            continue
        value = (dn.text or "").strip()
        if not value or "#REF" in value or re.search(r"\[\d+\]", value):
            continue
        local_sheet_id_raw = dn.get("localSheetId")
        if local_sheet_id_raw is not None:
            if sheet_idx is not None and int(local_sheet_id_raw) == sheet_idx:
                local_match = value
        else:
            workbook_match = value

    return local_match or workbook_match


def read_status_choices(xlsx_path: Path, structure: dict) -> list[str]:
    """Read dropdown choices for the status column from xlsx data validations.

    Inspects the data_validations attached to the worksheet and tries to extract
    the allowed values for the status column.  Three formats are handled:
    - Inline list: formula1 == '"Oui,Non,Partiel,N/A"' → split on comma.
    - Sheet range: formula1 == 'Ref!$A$1:$A$4' → read cells from that range.
    - Named range: formula1 == 'Cotation' → resolved via raw XML (handles the
      openpyxl #REF! shadow bug for local defined names).

    Args:
        xlsx_path: Path to the xlsx file.
        structure: Structure dict with col_status and selected_sheet keys.

    Returns:
        Ordered list of valid status strings, or [] if not found / not applicable.
    """
    col_status_letter = structure.get("col_status")
    if not col_status_letter:
        return []

    col_letter = col_status_letter.upper()

    try:
        wb = load_workbook(xlsx_path, data_only=True)
        sheet_name = structure.get("selected_sheet")
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        for dv in ws.data_validations.dataValidation:
            if dv.type != "list":
                continue

            # Check that this validation covers the status column
            sqref_str = str(dv.sqref).replace("$", "")
            if not re.search(rf'(?<![A-Z]){col_letter}\d', sqref_str):
                continue

            formula = (dv.formula1 or "").strip()
            if not formula:
                continue

            # Case 1 — inline list: "\"Oui,Non,Partiel,N/A\""
            if formula.startswith('"') and formula.endswith('"'):
                inner = formula[1:-1]
                choices = [c.strip() for c in inner.split(",") if c.strip()]
                if choices:
                    wb.close()
                    return choices

            # Case 2 — range reference: SheetName!$A$1:$A$4
            range_formula = formula.strip("'\"")
            try:
                if "!" in range_formula:
                    ref_sheet_name, cell_range = range_formula.split("!", 1)
                    ref_sheet_name = ref_sheet_name.strip("'")
                    if ref_sheet_name in wb.sheetnames:
                        ref_ws = wb[ref_sheet_name]
                        choices = [
                            str(row[0].value).strip()
                            for row in ref_ws[cell_range]
                            if row[0].value is not None and str(row[0].value).strip()
                        ]
                        if choices:
                            wb.close()
                            return choices
            except Exception:
                pass

            # Case 3 — named range: formula1 is a bare name like "Cotation"
            # openpyxl loses local defined names when a workbook-level #REF! entry
            # shadows them — resolve via raw XML instead.
            if re.match(r'^[A-Za-z_][A-Za-z0-9_.]*$', formula):
                sheet_idx = wb.sheetnames.index(ws.title) if ws.title in wb.sheetnames else None
                resolved = _resolve_named_range_from_xml(xlsx_path, formula, sheet_idx)
                if resolved and "!" in resolved:
                    try:
                        ref_sheet_name, cell_range = resolved.split("!", 1)
                        ref_sheet_name = ref_sheet_name.strip("'")
                        if ref_sheet_name in wb.sheetnames:
                            ref_ws = wb[ref_sheet_name]
                            choices = [
                                str(row[0].value).strip()
                                for row in ref_ws[cell_range]
                                if row[0].value is not None and str(row[0].value).strip()
                            ]
                            if choices:
                                wb.close()
                                return choices
                    except Exception:
                        pass

        wb.close()
    except Exception:
        logger.exception("read_status_choices: failed to read %s", xlsx_path.name)

    return []


def _write_cell(ws, row: int, col: int, value) -> None:
    """Write a value to a worksheet cell, handling merged cell ranges.

    In a merged range, only the top-left cell is writable. If the target cell
    is a MergedCell (read-only slave), find the merge range that contains it
    and write to its top-left corner instead.
    """
    from openpyxl.cell import MergedCell

    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        cell.value = value
        return

    # Find the merge range that owns this cell
    for merge_range in ws.merged_cells.ranges:
        if (merge_range.min_row <= row <= merge_range.max_row
                and merge_range.min_col <= col <= merge_range.max_col):
            ws.cell(row=merge_range.min_row, column=merge_range.min_col).value = value
            return

    # Fallback: should not happen, but skip silently rather than crash
    logger.warning("_write_cell: cell(%d,%d) is MergedCell but no range found — skipped", row, col)


def write_responses(
    source_path: Path,
    dest_path: Path,
    structure: dict,
    responses: list[dict],
) -> None:
    """Copy source_path to dest_path and insert Claude responses into col_response.

    Preserves Excel dropdown definitions (local defined names) using the same
    safe_local_defined_names pattern as the roundtrip endpoint.

    Args:
        source_path: Path to the anonymized xlsx (never modified — R-01).
        dest_path: Path to save the output xlsx with responses inserted.
        structure: Structure dict with col_response, selected_sheet, first_data_row,
                   col_id (optional).
        responses: List of {"question_id": str, "response": str} dicts.
    """
    col_r_letter = structure.get("col_response")
    if not col_r_letter:
        raise ValueError("Structure manquante : col_response est obligatoire.")

    col_r_idx = column_index_from_string(col_r_letter.upper())  # 1-based openpyxl
    col_id_letter = structure.get("col_id")
    col_id_idx = column_index_from_string(col_id_letter.upper()) if col_id_letter else None
    col_status_letter = structure.get("col_status")
    col_status_idx = column_index_from_string(col_status_letter.upper()) if col_status_letter else None
    first_data_row = int(structure.get("first_data_row") or 2)
    sheet_name = structure.get("selected_sheet")

    # Copy without modifying source (R-01)
    shutil.copy2(source_path, dest_path)

    # Preserve local defined names before openpyxl potentially clobbers them
    safe_names = safe_local_defined_names(dest_path)

    wb = load_workbook(dest_path, keep_links=False, rich_text=False)

    # Clear all defined names and re-inject only the safe local ones
    wb.defined_names.clear()
    for ws in wb.worksheets:
        ws.defined_names.clear()
        ws.print_area = None
        ws.print_title_rows = None
        ws.print_title_cols = None

    for name, attr_text, local_sheet_id in safe_names:
        dn = DefinedName(name=name, attr_text=attr_text, localSheetId=local_sheet_id)
        wb.defined_names.add(dn)

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # Build lookups: question_id → response text / status value
    resp_map: dict[str, str] = {r["question_id"]: r["response"] for r in responses}
    status_map: dict[str, str] = {
        r["question_id"]: r["status"] for r in responses if r.get("status")
    }

    # Pre-compute section-header rows: response column is merged into a column to its left.
    # Writing into such rows would overwrite the section header text — skip them entirely.
    header_rows: set[int] = _merged_header_rows(ws, col_r_idx)
    if header_rows:
        logger.info(
            "write_responses: %d section-header rows will be protected from overwrite: %s",
            len(header_rows),
            sorted(header_rows),
        )

    written = 0
    for row_idx in range(first_data_row, ws.max_row + 1):
        # Safeguard: never write into section-header rows (merged label rows)
        if row_idx in header_rows:
            continue

        # Get question_id for this row
        if col_id_idx is not None:
            id_cell = ws.cell(row=row_idx, column=col_id_idx)
            q_id = str(id_cell.value).strip() if id_cell.value is not None else str(row_idx)
        else:
            q_id = str(row_idx)

        if q_id in resp_map:
            _write_cell(ws, row_idx, col_r_idx, resp_map[q_id])
            written += 1

        if col_status_idx is not None and q_id in status_map:
            _write_cell(ws, row_idx, col_status_idx, status_map[q_id])

    wb.save(dest_path)
    wb.close()
    logger.info("write_responses: wrote %d responses into %s", written, dest_path.name)
