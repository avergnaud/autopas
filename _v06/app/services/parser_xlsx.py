"""Parsing and writing xlsx questionnaires with openpyxl."""
from __future__ import annotations

import logging
import re
import shutil
import zipfile
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string

from app.models.project import DocumentStructure
from app.services.anonymizer import Anonymizer

logger = logging.getLogger(__name__)
_MAX_EMPTY_ROWS = 20  # stop scanning after this many consecutive empty question cells


def extract_raw_content(filepath: Path, max_rows: int = 25) -> str:
    """Extract first max_rows rows from each sheet as text for Claude structure analysis."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    lines = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"=== Onglet : {sheet_name} ===")
        for i, row in enumerate(ws.iter_rows(max_row=max_rows, values_only=True), 1):
            cells = [str(c) if c is not None else "" for c in row]
            # Skip completely empty rows
            if any(c.strip() for c in cells):
                lines.append(f"Ligne {i} : {' | '.join(cells)}")
        lines.append("")
    wb.close()
    return "\n".join(lines)


def extract_questions(filepath: Path, structure: DocumentStructure) -> list[dict]:
    """Extract all questions from the xlsx. Returns list of {id, question, sheet, row}."""
    if not structure.sheets:
        return []

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    questions = []

    for sheet_info in structure.sheets:
        if not sheet_info.has_questions:
            continue
        if sheet_info.name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found in workbook", sheet_info.name)
            continue

        ws = wb[sheet_info.name]
        q_col_idx = column_index_from_string(sheet_info.question_column)
        id_col_idx = (
            column_index_from_string(sheet_info.id_column)
            if sheet_info.id_column
            else None
        )

        # Use iter_rows() — the correct streaming API for read_only mode.
        # ws.cell() + range(max_row) is O(n²) in read_only and can loop over
        # 1 048 576 rows if the xlsx dimension tag is stale.
        empty_streak = 0
        for row_idx, row_cells in enumerate(
            ws.iter_rows(min_row=sheet_info.first_data_row)
        ):
            row_num = sheet_info.first_data_row + row_idx
            q_val = row_cells[q_col_idx - 1].value if q_col_idx - 1 < len(row_cells) else None
            q_text = str(q_val).strip() if q_val is not None else ""

            if not q_text:
                empty_streak += 1
                if empty_streak >= _MAX_EMPTY_ROWS:
                    break
                continue
            empty_streak = 0

            q_id = ""
            if id_col_idx and id_col_idx - 1 < len(row_cells):
                id_val = row_cells[id_col_idx - 1].value
                q_id = str(id_val).strip() if id_val is not None else ""
            if not q_id:
                q_id = f"{sheet_info.name}_row{row_num}"

            questions.append({
                "id": q_id,
                "question": q_text,
                "sheet": sheet_info.name,
                "row": row_num,
            })

    wb.close()
    return questions


def format_questionnaire_for_claude(questions: list[dict]) -> str:
    """Format extracted questions as text block for the Claude prompt."""
    lines = []
    for q in questions:
        lines.append(f"ID: {q['id']}")
        lines.append(f"Question: {q['question']}")
        lines.append("")
    return "\n".join(lines)


def write_responses(
    filepath: Path, responses: list[dict], structure: DocumentStructure
) -> None:
    """Write Claude responses into response columns of the xlsx file."""
    if not structure.sheets:
        return

    response_map = {r["question_id"]: r["response"] for r in responses}

    wb = openpyxl.load_workbook(filepath)

    for sheet_info in structure.sheets:
        if not sheet_info.has_questions or sheet_info.name not in wb.sheetnames:
            continue

        ws = wb[sheet_info.name]
        q_col_idx = column_index_from_string(sheet_info.question_column)
        id_col_idx = (
            column_index_from_string(sheet_info.id_column)
            if sheet_info.id_column
            else None
        )
        resp_col_idxs = [
            column_index_from_string(c) for c in sheet_info.response_columns
        ]

        if not resp_col_idxs:
            logger.warning("No response columns for sheet '%s'", sheet_info.name)
            continue

        # Use iter_rows() to avoid relying on ws.max_row, which can be unreliable
        # for xlsx files with stale <dimension> tags (may return 1 048 576).
        empty_streak = 0
        for row_idx, row_cells in enumerate(ws.iter_rows(min_row=sheet_info.first_data_row)):
            row_num = sheet_info.first_data_row + row_idx

            q_val = row_cells[q_col_idx - 1].value if q_col_idx - 1 < len(row_cells) else None
            q_text = str(q_val).strip() if q_val is not None else ""

            # Track consecutive empty question cells to stop early
            if not q_text:
                empty_streak += 1
                if empty_streak >= _MAX_EMPTY_ROWS:
                    break
                continue
            empty_streak = 0

            # Determine question ID
            q_id = ""
            if id_col_idx and id_col_idx - 1 < len(row_cells):
                id_val = row_cells[id_col_idx - 1].value
                q_id = str(id_val).strip() if id_val is not None else ""
            if not q_id:
                q_id = f"{sheet_info.name}_row{row_num}"

            if q_id in response_map:
                ws.cell(row=row_num, column=resp_col_idxs[0]).value = response_map[q_id]

    wb.save(filepath)
    wb.close()
    fix_openpyxl_xml(filepath)


def delete_unlisted_sheets(filepath: Path, sheet_names_to_keep: list[str]) -> None:
    """Delete from the xlsx file every sheet whose name is NOT in sheet_names_to_keep."""
    if not sheet_names_to_keep:
        return
    wb = openpyxl.load_workbook(filepath)
    to_delete = [name for name in wb.sheetnames if name not in sheet_names_to_keep]
    for name in to_delete:
        logger.info("Deleting sheet '%s' from %s", name, filepath.name)
        del wb[name]
    if to_delete:
        wb.save(filepath)
    wb.close()


def apply_anonymization(filepath: Path, anonymizer: Anonymizer) -> None:
    """Apply anonymization to xlsx by modifying ZIP/XML content directly (memory-efficient)."""
    _xlsx_zip_replace(filepath, _make_xml_anonymizer(anonymizer))


def apply_deanonymization(filepath: Path, anonymizer: Anonymizer) -> None:
    """Apply deanonymization to xlsx by modifying ZIP/XML content directly (memory-efficient).

    Uses XML-escaped replacement values so that real names containing & < >
    produce valid XML (e.g., & → &amp;) instead of corrupting the file.
    """
    _xlsx_zip_replace(filepath, _make_xml_deanonymizer(anonymizer))


def fix_openpyxl_xml(filepath: Path) -> None:
    """Fix known openpyxl serialization bugs in the saved xlsx.

    Bug 1: openpyxl 3.x writes ``sqref=`` (lowercase) inside ``<dataValidation>``
    elements.  The OOXML spec (ECMA-376 §18.3.1.32) requires ``sqRef=``
    (camelCase).  Excel is case-sensitive and treats the lowercase variant as a
    missing attribute, which triggers its "We found a problem" repair dialog.

    Note: ``<selection sqref="…">`` is *correctly* lowercase per the spec
    (``CT_Selection`` §18.3.1.76) and must NOT be changed.

    Bug 2: openpyxl may drop the traditional ``externalLinkPath`` relationship
    (rId1) and only keep the newer ``externalLinkLongPath`` relationship under
    a different ID (e.g., rId2).  The ``externalLink{N}.xml`` file still
    references rId1, causing a broken-reference that triggers the repair dialog.

    Bug 3: openpyxl rewrites VML comment drawing files using generic namespace
    prefixes (ns0:, ns1:, …) and reuses the same prefix for different namespace
    URIs in different element scopes.  Excel's VML parser expects the conventional
    prefixes (v:, o:, x:) and may trigger the repair dialog when they are absent.

    Bug 4: when the original workbook has both an ``externalLinkPath`` and an
    ``externalLinkLongPath`` relationship for the same external book, openpyxl
    may retain only the ``externalLinkLongPath`` one (a Microsoft-proprietary
    2019+ type).  After ``_fix_extlink_rels`` renames its Id to match what
    ``externalLink{N}.xml`` expects, the relationship still has the wrong type
    URI.  Excel's ``<externalBook r:id="…">`` element expects a standard
    ``externalLinkPath`` relationship; when it finds ``externalLinkLongPath``
    instead it triggers the repair dialog.  Fix: downgrade the type to
    ``externalLinkPath`` for any rels file that contains *only*
    ``externalLinkLongPath`` entries (i.e. no ``externalLinkPath`` entry exists).
    """
    def _transform(text: str) -> str:
        def _fix_tag(m: re.Match) -> str:
            return m.group().replace("sqref=", "sqRef=")
        return re.sub(r"<dataValidation\b[^>]*>", _fix_tag, text)

    _xlsx_zip_replace(filepath, _transform)
    _fix_extlink_rels(filepath)
    _fix_extlink_longpath(filepath)
    _fix_vml_namespaces(filepath)


def _fix_extlink_rels(filepath: Path) -> None:
    """Fix relationship ID mismatches in external link rels files.

    When openpyxl saves a workbook that has both a traditional ``externalLinkPath``
    (rId1) and a modern ``externalLinkLongPath`` (rId2) per external link, it can
    drop the rId1 entry and keep only rId2.  The externalLink XML still references
    rId1, so Excel detects the orphaned reference and shows its repair dialog.

    Fix: for each external link where the rels file has exactly one entry whose ID
    does NOT match what the XML expects, rename that entry's ID to match.
    """
    id_renames: dict[str, dict[str, str]] = {}

    with zipfile.ZipFile(filepath, "r") as zin:
        nameset = set(zin.namelist())
        for name in sorted(nameset):
            m = re.match(r"xl/externalLinks/externalLink(\d+)\.xml$", name)
            if not m:
                continue
            rels_path = f"xl/externalLinks/_rels/externalLink{m.group(1)}.xml.rels"
            if rels_path not in nameset:
                continue
            xml = zin.read(name).decode("utf-8")
            rels = zin.read(rels_path).decode("utf-8")
            ref = re.search(r'<externalBook\b[^>]+\br:id="([^"]+)"', xml)
            if not ref:
                continue
            expected_id = ref.group(1)
            actual_ids = re.findall(r'\bId="([^"]+)"', rels)
            if expected_id not in actual_ids and len(actual_ids) == 1:
                id_renames[rels_path] = {actual_ids[0]: expected_id}
                logger.info(
                    "fix_extlink_rels: %s — renaming Id %r → %r",
                    rels_path, actual_ids[0], expected_id,
                )

    if not id_renames:
        return

    tmp = filepath.with_suffix(".~tmp.xlsx")
    try:
        with (
            zipfile.ZipFile(filepath, "r") as zin,
            zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as zout,
        ):
            for info in zin.infolist():
                data = zin.read(info.filename)
                if info.filename in id_renames:
                    try:
                        text = data.decode("utf-8")
                        for old_id, new_id in id_renames[info.filename].items():
                            text = text.replace(f'Id="{old_id}"', f'Id="{new_id}"')
                        data = text.encode("utf-8")
                    except Exception as exc:
                        logger.warning("fix_extlink_rels failed for %s: %s", info.filename, exc)
                zout.writestr(info, data)
        shutil.move(str(tmp), str(filepath))
    except Exception:
        try:
            tmp.unlink(missing_ok=True)
        except Exception:
            pass
        raise


def _fix_extlink_longpath(filepath: Path) -> None:
    """Downgrade externalLinkLongPath-only rels to externalLinkPath.

    When openpyxl saves a workbook that has both an ``externalLinkPath``
    (standard, file:// or https://) and an ``externalLinkLongPath``
    (Microsoft-proprietary 2019+ type) for the same external book, it may drop
    the ``externalLinkPath`` entry and retain only ``externalLinkLongPath``.
    After ``_fix_extlink_rels`` has corrected the Id, the type URI is still
    wrong: ``<externalBook r:id="rId1">`` expects ``externalLinkPath`` but finds
    ``externalLinkLongPath``, which triggers the Excel repair dialog.

    Fix: for each externalLink rels file that contains *only*
    ``externalLinkLongPath`` relationships (no ``externalLinkPath`` present),
    replace the type URI so that it becomes a standard ``externalLinkPath``.
    """
    _LONG_PATH_TYPE = (
        "http://schemas.microsoft.com/office/2019/04/relationships/externalLinkLongPath"
    )
    _STD_PATH_TYPE = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLinkPath"
    )

    files_to_fix: list[str] = []

    with zipfile.ZipFile(filepath, "r") as zin:
        nameset = set(zin.namelist())
        for name in sorted(nameset):
            if not re.match(r"xl/externalLinks/_rels/externalLink\d+\.xml\.rels$", name):
                continue
            text = zin.read(name).decode("utf-8")
            has_long = _LONG_PATH_TYPE in text
            has_std = _STD_PATH_TYPE in text
            if has_long and not has_std:
                files_to_fix.append(name)
                logger.info("fix_extlink_longpath: %s — downgrading externalLinkLongPath → externalLinkPath", name)

    if not files_to_fix:
        return

    tmp = filepath.with_suffix(".~tmp.xlsx")
    try:
        with (
            zipfile.ZipFile(filepath, "r") as zin,
            zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as zout,
        ):
            for info in zin.infolist():
                data = zin.read(info.filename)
                if info.filename in files_to_fix:
                    try:
                        text = data.decode("utf-8")
                        text = text.replace(_LONG_PATH_TYPE, _STD_PATH_TYPE)
                        data = text.encode("utf-8")
                    except Exception as exc:
                        logger.warning("fix_extlink_longpath failed for %s: %s", info.filename, exc)
                zout.writestr(info, data)
        shutil.move(str(tmp), str(filepath))
    except Exception:
        try:
            tmp.unlink(missing_ok=True)
        except Exception:
            pass
        raise


def _fix_vml_namespaces(filepath: Path) -> None:
    """Normalize VML namespace prefixes to standard Excel conventions.

    openpyxl rewrites VML files (xl/drawings/*.vml) using generic namespace
    prefixes (ns0:, ns1:, …) and reuses the same prefix for different namespace
    URIs in different element scopes (e.g. ns0 → office namespace in one element,
    then ns0 → vml namespace in a sibling).  Excel's VML parser is sensitive to
    this inconsistency and may trigger the "We found a problem" repair dialog.

    Fix: re-parse each .vml file with ElementTree after registering standard
    prefixes, then re-serialize so that all namespace declarations are hoisted
    to the root element with consistent, conventional prefixes::

        v:  urn:schemas-microsoft-com:vml
        o:  urn:schemas-microsoft-com:office:office
        x:  urn:schemas-microsoft-com:office:excel
    """
    import xml.etree.ElementTree as ET

    _VML_NAMESPACES = {
        "v": "urn:schemas-microsoft-com:vml",
        "o": "urn:schemas-microsoft-com:office:office",
        "x": "urn:schemas-microsoft-com:office:excel",
    }
    for prefix, uri in _VML_NAMESPACES.items():
        ET.register_namespace(prefix, uri)

    def _normalize(text: str) -> str:
        root = ET.fromstring(text)
        return ET.tostring(root, encoding="unicode")

    with zipfile.ZipFile(filepath, "r") as zin:
        vml_files = [n for n in zin.namelist() if n.endswith(".vml")]

    if not vml_files:
        return

    tmp = filepath.with_suffix(".~tmp.xlsx")
    try:
        with (
            zipfile.ZipFile(filepath, "r") as zin,
            zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as zout,
        ):
            for info in zin.infolist():
                data = zin.read(info.filename)
                if info.filename.endswith(".vml"):
                    try:
                        data = _normalize(data.decode("utf-8")).encode("utf-8")
                        logger.info("fix_vml_namespaces: normalized %s", info.filename)
                    except Exception as exc:
                        logger.warning("fix_vml_namespaces failed for %s: %s", info.filename, exc)
                zout.writestr(info, data)
        shutil.move(str(tmp), str(filepath))
    except Exception:
        try:
            tmp.unlink(missing_ok=True)
        except Exception:
            pass
        raise


def _make_xml_anonymizer(anonymizer: Anonymizer):
    """Return a transform that anonymizes text inside XML, handling XML-encoded chars.

    In xlsx XML, a cell containing "A & B" is stored as "A &amp; B".
    The plain anonymizer won't find the raw "&" form, so we also search for
    the XML-encoded variant of each key.
    """
    import xml.sax.saxutils as _sx

    def transform(text: str) -> str:
        for key in anonymizer.sorted_keys:
            alias = anonymizer.mappings[key]
            # Match raw form (case-insensitive)
            pattern = re.compile(re.escape(key), re.IGNORECASE)
            text = pattern.sub(alias, text)
            # Also match XML-encoded form when the key contains special chars
            xml_key = _sx.escape(key)
            if xml_key != key:
                pattern_xml = re.compile(re.escape(xml_key), re.IGNORECASE)
                text = pattern_xml.sub(alias, text)
        return text

    return transform


def _make_xml_deanonymizer(anonymizer: Anonymizer):
    """Return a transform that deanonymizes text inside XML with safe escaping.

    Real names are XML-escaped before insertion so that names containing
    & < > produce valid XML (e.g., & → &amp;) rather than corrupting it.
    """
    import xml.sax.saxutils as _sx

    reverse = {v: k for k, v in anonymizer.mappings.items()}
    sorted_keys = sorted(reverse.keys(), key=len, reverse=True)

    def transform(text: str) -> str:
        for key in sorted_keys:
            real_name = reverse[key]
            xml_safe_name = _sx.escape(real_name)
            text = text.replace(key, xml_safe_name)
        return text

    return transform


def _xlsx_zip_replace(filepath: Path, transform_fn) -> None:
    """Memory-efficient text replacement inside an xlsx (ZIP) by processing XML entries.

    Targets xl/sharedStrings.xml (where most text is stored) and worksheet XML files
    (for inline strings). Avoids loading a full openpyxl workbook in write mode.
    """
    tmp = filepath.with_suffix(".~tmp.xlsx")
    try:
        with (
            zipfile.ZipFile(filepath, "r") as zin,
            zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as zout,
        ):
            for info in zin.infolist():
                data = zin.read(info.filename)
                name = info.filename
                if name == "xl/sharedStrings.xml" or (
                    name.startswith("xl/worksheets/") and name.endswith(".xml")
                ):
                    try:
                        text = data.decode("utf-8")
                        text = transform_fn(text)
                        data = text.encode("utf-8")
                    except Exception as exc:
                        logger.warning("ZIP replace failed for %s: %s", name, exc)
                zout.writestr(info, data)
        shutil.move(str(tmp), str(filepath))
    except Exception:
        try:
            tmp.unlink(missing_ok=True)
        except Exception:
            pass
        raise
