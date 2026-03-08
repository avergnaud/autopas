"""PAS Assistant — xlsx questionnaire parser.

Handles reading questions from an anonymized xlsx and writing Claude-generated
responses back into a copy of that file.
"""

import logging
import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.workbook.defined_name import DefinedName

from app.services.anonymizer import safe_local_defined_names

logger = logging.getLogger(__name__)


def read_questions(xlsx_path: Path, structure: dict) -> list[dict]:
    """Extract questions from an xlsx file using confirmed structure.

    Args:
        xlsx_path: Path to the xlsx file (anonymized).
        structure: Structure dict with keys: selected_sheet, first_data_row,
                   col_id (optional), col_question.

    Returns:
        List of dicts with keys: row (int), question_id (str), question_text (str).
        Rows where question_text is empty are skipped.
    """
    sheet_name = structure.get("selected_sheet")
    first_data_row = int(structure.get("first_data_row") or 2)
    col_id_letter = structure.get("col_id")
    col_q_letter = structure.get("col_question")

    if not col_q_letter:
        raise ValueError("Structure manquante : col_question est obligatoire.")

    col_q_idx = column_index_from_string(col_q_letter.upper())  # 1-based
    col_id_idx = column_index_from_string(col_id_letter.upper()) if col_id_letter else None

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    questions: list[dict] = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx < first_data_row:
            continue

        # Extract question text
        q_text_raw = row[col_q_idx - 1] if col_q_idx - 1 < len(row) else None
        q_text = str(q_text_raw).strip() if q_text_raw is not None else ""
        if not q_text:
            continue

        # Extract question id
        if col_id_idx is not None and col_id_idx - 1 < len(row):
            q_id_raw = row[col_id_idx - 1]
            q_id = str(q_id_raw).strip() if q_id_raw is not None else str(row_idx)
        else:
            q_id = str(row_idx)

        questions.append({"row": row_idx, "question_id": q_id, "question_text": q_text})

    wb.close()
    logger.info("read_questions: extracted %d questions from %s", len(questions), xlsx_path.name)
    return questions


def write_responses(
    source_path: Path,
    dest_path: Path,
    structure: dict,
    responses: list[dict],
) -> None:
    """Copy source_path to dest_path and insert Claude responses into col_response.

    Preserves Excel dropdown definitions (local defined names) using the same
    safe_local_defined_names pattern as the roundtrip endpoint.

    Args:
        source_path: Path to the anonymized xlsx (never modified — R-01).
        dest_path: Path to save the output xlsx with responses inserted.
        structure: Structure dict with col_response, selected_sheet, first_data_row,
                   col_id (optional).
        responses: List of {"question_id": str, "response": str} dicts.
    """
    col_r_letter = structure.get("col_response")
    if not col_r_letter:
        raise ValueError("Structure manquante : col_response est obligatoire.")

    col_r_idx = column_index_from_string(col_r_letter.upper())  # 1-based openpyxl
    col_id_letter = structure.get("col_id")
    col_id_idx = column_index_from_string(col_id_letter.upper()) if col_id_letter else None
    first_data_row = int(structure.get("first_data_row") or 2)
    sheet_name = structure.get("selected_sheet")

    # Copy without modifying source (R-01)
    shutil.copy2(source_path, dest_path)

    # Preserve local defined names before openpyxl potentially clobbers them
    safe_names = safe_local_defined_names(dest_path)

    wb = load_workbook(dest_path, keep_links=False, rich_text=False)

    # Clear all defined names and re-inject only the safe local ones
    wb.defined_names.clear()
    for ws in wb.worksheets:
        ws.defined_names.clear()
        ws.print_area = None
        ws.print_title_rows = None
        ws.print_title_cols = None

    for name, attr_text, local_sheet_id in safe_names:
        dn = DefinedName(name=name, attr_text=attr_text, localSheetId=local_sheet_id)
        wb.defined_names.add(dn)

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # Build lookup: question_id → response text
    resp_map: dict[str, str] = {r["question_id"]: r["response"] for r in responses}

    written = 0
    for row_idx in range(first_data_row, ws.max_row + 1):
        # Get question_id for this row
        if col_id_idx is not None:
            id_cell = ws.cell(row=row_idx, column=col_id_idx)
            q_id = str(id_cell.value).strip() if id_cell.value is not None else str(row_idx)
        else:
            q_id = str(row_idx)

        if q_id in resp_map:
            ws.cell(row=row_idx, column=col_r_idx).value = resp_map[q_id]
            written += 1

    wb.save(dest_path)
    wb.close()
    logger.info("write_responses: wrote %d responses into %s", written, dest_path.name)
