"""Structure detection for questionnaire xlsx files using Claude API.

Reads the anonymized xlsx, extracts a column-labeled preview of each sheet,
then asks Claude to identify: selected sheet, header row, first data row,
ID column, question column, response column, and optional status column.
"""

import json
import logging
import os
import string
from pathlib import Path

import anthropic
from openpyxl import load_workbook

from app.config import BASE_DIR

logger = logging.getLogger(__name__)

_MAX_PREVIEW_ROWS = 30
_MAX_PREVIEW_COLS = 20
_PROMPTS_DIR = BASE_DIR / "data" / "config" / "prompts"


def _load_prompt(name: str) -> str:
    """Load a prompt file from data/config/prompts/.

    Args:
        name: Filename within the prompts directory.

    Returns:
        File contents as a string.

    Raises:
        RuntimeError: If the file does not exist.
    """
    path = _PROMPTS_DIR / name
    if not path.exists():
        raise RuntimeError(
            f"Fichier prompt introuvable : {path}. "
            "Vérifiez que data/config/prompts/ est correctement déployé."
        )
    return path.read_text(encoding="utf-8")


def _col_letter(idx: int) -> str:
    """Convert 0-based column index to Excel letter (0→A, 25→Z, 26→AA…)."""
    result = ""
    idx += 1
    while idx:
        idx, rem = divmod(idx - 1, 26)
        result = string.ascii_uppercase[rem] + result
    return result


def _extract_preview(xlsx_path: Path) -> dict[str, list[list[str]]]:
    """Read the first N rows of each sheet as lists of strings."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheets: dict[str, list[list[str]]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows: list[list[str]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= _MAX_PREVIEW_ROWS:
                break
            cells = [str(c) if c is not None else "" for c in row[:_MAX_PREVIEW_COLS]]
            rows.append(cells)
        sheets[name] = rows
    wb.close()
    return sheets


def _format_preview(sheets: dict[str, list[list[str]]]) -> str:
    """Format sheets as a readable text table with column letters."""
    lines: list[str] = []
    for sheet_name, rows in sheets.items():
        if not any(any(c for c in row) for row in rows):
            continue
        # Determine actual number of columns in this sheet
        ncols = max((len(r) for r in rows), default=0)
        col_header = "      | " + " | ".join(f"{_col_letter(i):^12}" for i in range(ncols))

        lines.append(f"\n=== Onglet : {sheet_name} ===")
        lines.append(col_header)
        lines.append("      | " + "-" * (14 * ncols))

        for i, row in enumerate(rows):
            padded = (row + [""] * ncols)[:ncols]
            # Truncate long cell values for readability
            cells = [c[:12].ljust(12) for c in padded]
            lines.append(f"  L{i + 1:02d}  | " + " | ".join(cells))

    return "\n".join(lines)


def detect_xlsx_structure(xlsx_path: Path) -> dict:
    """Call Claude to detect the questionnaire structure in an anonymized xlsx.

    Args:
        xlsx_path: Path to the anonymized xlsx file.

    Returns:
        Dict with keys: selected_sheet, all_sheets, header_row, first_data_row,
        col_id, col_question, col_response, col_status, confidence, notes.

    Raises:
        RuntimeError: If ANTHROPIC_API_KEY is not set or Claude returns invalid JSON.
    """
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key or api_key.startswith("sk-ant-..."):
        raise RuntimeError("ANTHROPIC_API_KEY non configurée.")

    sheets = _extract_preview(xlsx_path)
    preview = _format_preview(sheets)
    all_sheets = list(sheets.keys())

    system = (
        "Tu es un expert en analyse de fichiers Excel de questionnaires de sécurité. "
        "Réponds UNIQUEMENT en JSON valide, sans texte avant ou après."
    )

    user = _load_prompt("system_structure.txt").replace("{preview}", preview)

    model = os.environ.get("CLAUDE_MODEL", "claude-sonnet-4-6")
    client = anthropic.Anthropic(api_key=api_key)
    response = client.messages.create(
        model=model,
        max_tokens=512,
        system=system,
        messages=[{"role": "user", "content": user}],
    )

    raw = response.content[0].text.strip()
    # Strip markdown code block if present
    if raw.startswith("```"):
        parts = raw.split("```")
        raw = parts[1] if len(parts) > 1 else raw
        if raw.lower().startswith("json"):
            raw = raw[4:]
    raw = raw.strip()

    result: dict = json.loads(raw)
    result["all_sheets"] = all_sheets  # always use real sheet names

    logger.info(
        "Structure detected for %s: sheet=%s header=%s id=%s q=%s r=%s status=%s evidence=%s conf=%s",
        xlsx_path.name,
        result.get("selected_sheet"),
        result.get("header_row"),
        result.get("col_id"),
        result.get("col_question"),
        result.get("col_response"),
        result.get("col_status"),
        result.get("col_evidence"),
        result.get("confidence"),
    )
    return result
